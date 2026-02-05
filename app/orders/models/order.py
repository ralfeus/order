''' Order model '''
from __future__ import annotations
from datetime import datetime
from decimal import Decimal
import logging
from functools import reduce
import os.path
from tempfile import _TemporaryFileWrapper, NamedTemporaryFile
from typing import Optional
import openpyxl
from openpyxl.styles import PatternFill

from sqlalchemy import Boolean, Column, Enum, DateTime, Numeric, ForeignKey, Integer, \
    String, func, or_
from sqlalchemy.ext.associationproxy import association_proxy
from sqlalchemy.orm import backref, relationship
from sqlalchemy.orm.attributes import InstrumentedAttribute
from sqlalchemy.orm.collections import attribute_mapped_collection

from app import db
from app.currencies.models.currency import Currency
import app.invoices.models as i
import app.orders.models.suborder as so
import app.payments.models.payment_method as pm
from app.models.base import BaseModel
from app.models.country import Country
from app.orders.signals import sale_order_model_preparing
import app.purchase.models as p
from app.settings.models.setting import Setting
from app.shipping.models.shipping import Shipping
from app.users.models.user import User
from exceptions import OrderError, UnfinishedOrderError

from .order_product import OrderProduct
from .order_status import OrderStatus

ORDER_ID = 'orders.id'

class OrderBox(db.Model, BaseModel): # type: ignore
    ''' Specific box used in order '''
    __tablename__ = 'order_boxes'

    order_id: str = Column(String(16), ForeignKey(ORDER_ID))
    length: int = Column(Integer)
    width: int = Column(Integer)
    height: int = Column(Integer)
    weight: int = Column(Integer)
    quantity: int = Column(Integer)

    def to_dict(self):
        return {
            'length': self.length,
            'width': self.width,
            'height': self.height,
            'weight': self.weight,
            'quantity': self.quantity
        }

class Order(db.Model, BaseModel): # type: ignore
    ''' Sale order '''
    __tablename__ = 'orders'
    __id_pattern = 'ORD-{year}-{month:02d}-'
    service_fee = 0

    id = Column(String(16), primary_key=True, nullable=False)
    seq_num = Column(Integer)
    user_id = Column(Integer, ForeignKey('users.id'))
    user = relationship(User, foreign_keys=[user_id])
    invoice_id = Column(String(16), ForeignKey('invoices.id'))
    invoice: i.Invoice = relationship('Invoice', foreign_keys=[invoice_id]) # type: ignore
    customer_name = Column(String(64))
    address = Column(String(256))
    city_eng: str = Column(String(128)) # type: ignore
    country_id: str = Column(String(2), ForeignKey('countries.id')) # type: ignore
    country: Country = relationship(Country, foreign_keys=[country_id]) # type: ignore
    zip = Column(String(15))
    phone = Column(String(64))
    email = Column(String(64))
    comment = Column(String(65536))
    boxes: list[OrderBox] = relationship('OrderBox', lazy='dynamic', cascade="all, delete-orphan") # type: ignore
    shipping_box_weight = Column(Integer())
    total_weight = Column(Integer(), default=0)
    total_weight_set_manually = Column(Boolean, default=False)
    shipping_method_id = Column(Integer, ForeignKey('shipping.id'))
    shipping: Shipping = relationship('Shipping', foreign_keys=[shipping_method_id]) # type: ignore
    subtotal_krw = Column(Integer(), default=0)
    subtotal_cur1 = Column(Numeric(10, 2), default=0)
    subtotal_cur2 = Column(Numeric(10, 2), default=0)
    shipping_krw = Column(Integer(), default=0)
    shipping_cur1 = Column(Numeric(10, 2), default=0)
    shipping_cur2 = Column(Numeric(10, 2), default=0)
    total_krw = Column(Integer(), default=0)
    total_cur1 = Column(Numeric(10, 2), default=0)
    total_cur2 = Column(Numeric(10, 2), default=0)
    status = Column(Enum(OrderStatus), default='pending')
    tracking_id = Column(String(64))
    tracking_url:str = Column(String(256)) # type: ignore
    when_created = Column(DateTime)
    when_changed = Column(DateTime)
    purchase_date = Column(DateTime)
    purchase_date_sort = Column(DateTime, index=True,
        nullable=False, default=datetime(9999, 12, 31))
    suborders:list[so.Suborder] = relationship('Suborder', lazy='dynamic', cascade='all, delete-orphan') # type: ignore
    __order_products = relationship('OrderProduct', lazy='dynamic')
    attached_order_id = Column(String(16), ForeignKey(ORDER_ID))
    attached_order = relationship('Order', remote_side=[id])
    attached_orders = relationship('Order',
        foreign_keys=[attached_order_id], lazy='dynamic')
    payment_method_id = Column(Integer(), ForeignKey('payment_methods.id'))
    payment_method: pm.PaymentMethod = relationship('PaymentMethod', foreign_keys=[payment_method_id]) # type: ignore
    transaction_id = Column(Integer(), ForeignKey('transactions.id'))
    transaction = relationship('Transaction', foreign_keys=[transaction_id])
    params:dict[str, str] = association_proxy('order_params', 'value',
        creator=lambda k, v: OrderParam(name=k, value=v,
                                        when_created=datetime.now(),
                                        when_changed=datetime.now())
    )

    @property
    def order_products(self) -> list[OrderProduct]:
        ''' Returns aggregated list of order products for all suborders '''
        if self.suborders.count() > 0: #type: ignore
            return [order_product for suborder in self.suborders
                                  for order_product in suborder.order_products]
        return list(self.__order_products)

    @property
    def purchase_orders(self):
        ''' Returns list of purchase orders for all suborders '''
        from app.purchase.models.purchase_order import PurchaseOrder
        return PurchaseOrder.query.filter(
            PurchaseOrder.suborder_id.in_([suborder.id for suborder in self.suborders]))

    def set_purchase_date(self, value):
        self.purchase_date = value
        self.purchase_date_sort = value

    def set_status(self, value, actor):
        ''' Sets status of the order '''
        logger = logging.getLogger(f'orders.models.Order::set_status():{self.id}')
        if isinstance(value, str):
            value = OrderStatus[value.lower()]
        elif isinstance(value, int):
            value = OrderStatus(value)

        if value not in [OrderStatus.pending, OrderStatus.can_be_paid]:
            self.purchase_date_sort = datetime(9999, 12, 31)
        if value == OrderStatus.shipped:
            from app.orders.models.order_product import OrderProductStatus
            unfinished_ops = []
            for suborder in self.suborders:
                for order_product in suborder.order_products:
                    if order_product.status not in [OrderProductStatus.unavailable,
                                                    OrderProductStatus.purchased]:
                        unfinished_ops.append("{} - {}: {}".format(
                            suborder.id, order_product.product_id, order_product.status
                        ))
            if len(unfinished_ops) > 0:
                raise UnfinishedOrderError(unfinished_ops)
            for ao in self.attached_orders:
                ao.set_status(value, actor)
            self.__pay(actor)
        self.status = value
        if self.status == OrderStatus.shipped:
            from app.orders.signals import sale_order_shipped
            sale_order_shipped.send(self)

    @staticmethod
    def get_new_id():
        ''' Generates new seq_num and ID for the order '''
        today = datetime.now()
        today_prefix = Order.__id_pattern.format(year=today.year, month=today.month)
        last_order = db.session.query(Order.seq_num). \
            filter(Order.id.like(today_prefix + '%')). \
            order_by(Order.seq_num.desc()). \
            first()
        seq_num = last_order[0] + 1 if last_order else 1
        id = today_prefix + '{:04d}'.format(seq_num)
        return seq_num, id

    def __init__(self, **kwargs):
        super().__init__(**kwargs)
        self.seq_num, self.id = self.get_new_id()

        self.total_weight = 0
        self.total_krw = 0
        self.when_created = datetime.now()

        attributes = [a[0] for a in type(self).__dict__.items()
                           if isinstance(a[1], InstrumentedAttribute)]
        for arg in kwargs:
            if arg in attributes:
                setattr(self, arg, kwargs[arg])
        # Here properties are set (attributes start with '__')
        if kwargs.get('shipping'):
            self.shipping = kwargs['shipping']
        if kwargs.get('status'):
            self.status = kwargs['status']
        self.service_fee = kwargs.get('service_fee', 0)

    def __repr__(self):
        return f"<Order: {self.id}>"

    def attach_orders(self, orders):
        if isinstance(orders, list) and len(orders) > 0:
            if isinstance(orders[0], Order):
                self.attached_orders = orders
            else:
                self.attached_orders = Order.query.filter(Order.id.in_(orders))
        else:
            self.attached_orders = []

    def __pay(self, actor):
        logger = logging.getLogger('Order.__pay')
        from app.payments.models.transaction import Transaction
        #TODO: wrong approach
        if not self.total_krw:
            logger.debug("%s totals are undefined. Updating...", self.id)
            self.update_total()
        transaction = Transaction(
            amount=-self.total_krw,
            customer=self.user,
            user=actor
        )
        self.transaction = transaction
        db.session.add(transaction)

    def delete(self):
        for suborder in self.suborders:
            suborder.delete()
        OrderParam.query.filter_by(order_id=self.id).delete()
        super().delete()

    @classmethod
    def get_filter(cls, base_filter, column = None, filter_value = None):
        if filter_value is None:
            return base_filter
        from .suborder import Suborder
        from app.invoices.models.invoice import Invoice
        from app.purchase.models.purchase_order import PurchaseOrder
        from app.users.models.user import User
        part_filter = f'%{filter_value}%'
        filter_values = filter_value.split(',')
        if column is None:
            return \
                base_filter.filter(or_(
                    cls.id.like(part_filter),
                    cls.customer_name.like(part_filter),
                    cls.user.has(User.username.like(part_filter)),
                    cls.comment.like(part_filter),
                    cls.status.like(part_filter),
                    cls.tracking_id.like(part_filter)
                ))
        if isinstance(column, str):
            if column == 'when_po_posted':
                return base_filter.filter(
                    PurchaseOrder.query.filter(
                        PurchaseOrder.suborder_id == Suborder.id,
                        func.date(PurchaseOrder.when_posted) == filter_value,
                        Suborder.order_id == Order.id).exists())
            elif column == 'invoice_export_id':
                if filter_value == 'NULL':
                    return base_filter.filter(Order.invoice == None)
                else:
                    return base_filter.filter(
                        Order.invoice.has(Invoice.export_id.like(part_filter)))
            else:
                return base_filter
        return \
            base_filter.filter(cls.country_id.in_(filter_values)) \
                if column.key == 'country' else \
            base_filter.filter(Order.payment_method_id.in_(filter_values)) \
                if column.key == 'payment_method' else \
            base_filter.filter(Order.shipping_method_id.in_(filter_values)) \
                if column.key == 'shipping' else \
            base_filter.filter(column.has(User.username.like(part_filter))) \
                if column.key == 'user' else \
            base_filter.filter(column.in_([OrderStatus[status]
                               for status in filter_values])) \
                if column.key == 'status' \
            else base_filter.filter(column.like(part_filter))

    def get_boxes_weight(self):
        return reduce(lambda acc, b: acc + b.weight * b.quantity, self.boxes, 0)

    def get_payee(self) -> Optional[p.Company]:
        return self.payment_method.payee if self.payment_method \
            else None
    
    def get_service_fee(self, currency: Optional[Currency]=None):
        return (self.service_fee * currency.rate) if currency else self.service_fee
    
    def get_shipping(self, currency: Optional[Currency]=None):
        ''' Returns shipping cost in currency provided '''
        return \
            self.shipping_cur1 if currency and currency.code == 'USD' \
            else self.shipping_cur2 if currency and currency.code == 'EUR' \
            else self.shipping_krw if currency and currency.base \
            else round(self.shipping_krw * currency.rate, currency.decimal_places or 0) if currency \
            else self.shipping_krw

    def get_subtotal(self, currency: Optional[Currency]=None):
        '''Returns subtotal of the order - sum of cost of all order products'''
        return \
            self.subtotal_cur1 if currency and currency.code == 'USD' \
            else self.subtotal_cur2 if currency and currency.code == 'EUR' \
            else self.subtotal_krw if currency and currency.base \
            else round(self.subtotal_krw * currency.rate, currency.decimal_places or 0) if currency \
            else self.subtotal_krw

    def get_total(self, currency: Optional[Currency]=None):
        '''Returns total of the order - subtotal plus shipping costs plus service fee'''
        return \
            self.total_cur1 if currency and currency.code == 'USD' \
            else self.total_cur2 if currency and currency.code == 'EUR' \
            else self.total_krw if currency and currency.base \
            else round(self.total_krw * currency.rate, currency.decimal_places or 0) if currency \
            else self.total_krw

    def get_total_points(self):
        return reduce(
            lambda acc, sub: acc + sub.get_total_points(),
            self.suborders, 0)

    def is_editable(self):
        return self.status in [OrderStatus.draft]

    def need_to_update_total(self, payload):
        return len({'subtotal_krw', 'total_weight', 'shipping_krw', 'country', 'suborders'} &
                   set(payload.keys())) > 0

    def to_dict(self, details=False, partial=None):
        ''' Returns dictionary representation of the object ready to be JSONified '''
        def list_to_dict(key_list, value):
            if len(key_list) == 0:
                return value
            return {key_list[0]: list_to_dict(key_list[1:], value)}
        def merge(a, b, path=None):
            "merges b into a"
            if path is None: path = []
            for key in b:
                if key in a:
                    if isinstance(a[key], dict) and isinstance(b[key], dict):
                        merge(a[key], b[key], path + [str(key)])
                    elif a[key] == b[key]:
                        pass # same leaf value
                    else:
                        raise Exception('Conflict at %s' % '.'.join(path + [str(key)]))
                else:
                    a[key] = b[key]
            return a
        logger = logging.getLogger('Order.to_dict')
        from app.payments.models.payment import PaymentStatus
        from app.purchase.models.purchase_order import PurchaseOrder
        from .suborder import Suborder
        is_order_updated = False
        if not self.total_krw:
            logger.debug("%s totals are undefined. Updating...", self.id)
            self.update_total()
            is_order_updated = True
        if not self.total_cur1:
            logger.debug("%s total USD is undefined. Updating...", self.id)
            self.total_cur1 = self.total_krw * Currency.query.get('USD').rate
            is_order_updated = True
        if not self.total_cur2:
            logger.debug("%s total EUR is undefined. Updating...", self.id)
            curr_eur = Currency.query.get('EUR')
            rate = curr_eur.rate if curr_eur is not None else 0
            self.total_cur2 = self.total_krw * rate
            is_order_updated = True
        if is_order_updated:
            db.session.commit()
        check_outsiders_setting = Setting.query.get('check_outsiders')
        need_to_check_outsiders = check_outsiders_setting.value == '1' \
            if check_outsiders_setting is not None else False
        posted_pos = (
            PurchaseOrder.query.join(Suborder)
                .filter(Suborder.order_id == self.id)
                .filter(PurchaseOrder.when_posted != None)
        )
        when_po_posted = db.session.query(func.max(PurchaseOrder.when_posted)) \
            .select_entity_from(posted_pos.subquery()).scalar() \
            if posted_pos.count() == self.suborders.count() \
            else None

        # Issuing a signal to get module specific part of the order
        res = sale_order_model_preparing.send(self, details=details)
        ext_model = reduce(lambda acc, i: {**acc, **i}, [i[1] for i in res], {})

        result = {
            'id': self.id,
            'user': self.user.username if self.user else None,
            'customer_name': self.customer_name,
            'address': self.address,
            'city_eng': self.city_eng,
            'country': self.country.to_dict() if self.country else None,
            'zip': self.zip,
            'phone': self.phone,
            'email': self.email,
            'comment': self.comment,
            'invoice_id': self.invoice_id,
            'invoice': self.invoice.to_dict() if self.invoice is not None else None,
            'subtotal_krw': self.subtotal_krw,
            'total_weight': self.total_weight,
            'shipping_box_weight': self.shipping_box_weight,
            'shipping_krw': self.shipping_krw,
            'total': self.total_krw,
            'total_krw': self.total_krw,
            'total_cur1': float(self.total_cur1),
            'total_cur2': float(self.total_cur2),
            'shipping': self.shipping.to_dict() if self.shipping else None,
            'boxes': [box.to_dict() for box in self.boxes],
            'status': self.status.name if self.status else None,
            'payment_method': self.payment_method.name \
                if self.payment_method else None,
            'payment_pending': self.status == OrderStatus.pending \
                and self.payments.filter_by(status=PaymentStatus.pending).count() > 0,
            'tracking_id': self.tracking_id if self.tracking_id else None,
            'tracking_url': self.tracking_url if self.tracking_url else None,
            'outsiders': [so.subcustomer.username + ":" + so.subcustomer.name
                          for so in self.suborders
                          if not so.is_for_internal()] \
                         if need_to_check_outsiders \
                         else [],
            'params': reduce(
                lambda acc, pair: 
                    merge(acc, list_to_dict(pair[0].split('.'), pair[1])),
                self.params.items(), {}),
            'purchase_date': self.purchase_date.strftime('%Y-%m-%d %H:%M:%S') \
                if self.purchase_date else None,
            'when_po_posted': when_po_posted.strftime('%Y-%m-%d %H:%M:%S') \
                if when_po_posted else None,
            'when_created': self.when_created.strftime('%Y-%m-%d %H:%M:%S') \
                if self.when_created else None,
            'when_changed': self.when_changed.strftime('%Y-%m-%d %H:%M:%S') \
                if self.when_changed else None,
            **ext_model
        }
        if details:
            result = {**result,
                'suborders': [so.to_dict() for so in self.suborders],
                'order_products': [op.to_dict() for op in self.order_products],
                'attached_orders': [o.to_dict() for o in self.attached_orders],
                'purchase_orders': [po.to_dict() for po in self.purchase_orders]
            }
        if (partial is not None):
            result = {k: result.get(k) for k in partial}
        return result

    def update_total(self):
        ''' Updates totals of the order '''
        from app.shipping.models.shipping import PostponeShipping, Shipping, NoShipping
        logger = logging.getLogger(self.id)
        logger.debug("Updating total")
        logger.debug("There are %s suborders", self.suborders.count()) #type: ignore
        for suborder in self.suborders:
            suborder.update_total()
            logger.debug("The suborder %s:", suborder.id)
            logger.debug("\tLocal shipping (base): %s", suborder.local_shipping)
            logger.debug("\tSubtotal (base): %s", suborder.get_subtotal())
            logger.debug("\tTotal weight: %s", suborder.get_total_weight())

        order_weight = reduce(lambda acc, sub: acc + sub.get_total_weight(),
                              self.suborders, 0)
        logger.debug("Total order weight: %s", order_weight)
        attached_orders_weight = reduce(lambda acc, ao: acc + ao.total_weight,
                                        self.attached_orders, 0)
        logger.debug("Attached orders weight: %s", attached_orders_weight)
        if self.shipping is None:
            if self.shipping_method_id is not None:
                self.shipping = Shipping.query.get(self.shipping_method_id)
            else:
                self.shipping = NoShipping.query.first()
                if self.shipping is None:
                    self.shipping = NoShipping()
        logger.debug("Shipping: %s", self.shipping)
        if not self.total_weight_set_manually:
            self.total_weight = order_weight + attached_orders_weight
            self.shipping_box_weight = reduce(lambda acc, b: acc + b.weight,
                                            self.boxes, 0) \
                if self.boxes.count() > 0 \
                else self.shipping.get_box_weight(self.total_weight) \
                    if not isinstance(self.shipping, (NoShipping, PostponeShipping)) \
                    else 0
            logger.debug("Total weight (calculated): %s", self.total_weight)
            logger.debug("Box weight: %s", self.shipping_box_weight)
            logger.debug("Total weight (calculated) with box: %s",
                         self.total_weight + self.shipping_box_weight)
        else:
            self.shipping_box_weight = 0
            logger.debug("Total weight (set manually): %s", self.total_weight)
        # self.subtotal_krw = reduce(lambda acc, op: acc + op.price * op.quantity,
        #                            self.order_products, 0)
        self.subtotal_krw = reduce(
            lambda acc, sub: acc + sub.get_subtotal() + sub.local_shipping, self.suborders, 0)
        curr_eur = Currency.query.get('EUR')
        rate_eur = curr_eur.rate if curr_eur is not None else 0
        logger.debug("Subtotal: %s", self.subtotal_krw)
        self.subtotal_cur2 = self.subtotal_krw * float(rate_eur)
        self.subtotal_cur1 = self.subtotal_krw * float(Currency.query.get('USD').rate)

        self.shipping_krw = int(Decimal(self.shipping.get_shipping_cost(
            self.country.id if self.country else None,
            self.total_weight + self.shipping_box_weight)))
        logger.debug("Shipping (base): %s", self.shipping_krw)
        self.shipping_cur2 = self.shipping_krw * float(rate_eur)
        self.shipping_cur1 = self.shipping_krw * float(Currency.query.get('USD').rate)
        logger.debug("Service fee: %s", self.service_fee)
        self.total_krw = self.subtotal_krw + self.shipping_krw + self.service_fee
        logger.debug("Total (base): %s", self.total_krw)
        self.total_cur2 = self.subtotal_cur2 + self.shipping_cur2
        self.total_cur1 = self.subtotal_cur1 + self.shipping_cur1


    def get_order_excel(self) -> _TemporaryFileWrapper:
        '''Generates an invoice in excel format. Returns a temporary file object'''
        logger = logging.getLogger('Order.get_order_excel')
        if len(self.order_products) == 0:
            raise OrderError("The order has no products")
        if not self.total_krw:
            logger.debug("%s totals are undefined. Updating...", self.id)
            self.update_total()
        package_path = os.path.dirname(__file__) + '/..'
        suborder_fill = PatternFill(
            start_color='00FFFF00', end_color='00FFFF00', fill_type='solid')
        order_wb = openpyxl.open(f'{package_path}/templates/order_template.xlsx')
        ws = order_wb.worksheets[0]

        # Set order header
        ws.cell(2, 2, "\n".join([self.id] + [ao.id for ao in self.attached_orders]))
        ws.cell(2, 3, self.when_created.strftime('%Y-%m-%d'))
        ws.cell(4, 2, self.customer_name)
        ws.cell(5, 2, str(self.address) + '\n' + str(self.city_eng) + '\n' + str(self.zip))
        ws.cell(6, 2, self.phone)
        # Set currency rates
        ws.cell(8, 5, float(1 / Currency.query.get('EUR').rate))
        ws.cell(9, 5, float(1 / Currency.query.get('USD').rate))

        ws.cell(6, 6, self.subtotal_krw)
        ws.cell(6, 7, self.total_weight + self.shipping_box_weight)
        ws.cell(6, 8, self.shipping_krw)
        ws.cell(6, 9, self.total_krw)
        ws.cell(6, 13, reduce(lambda acc, op: acc + op.product.points,
                              self.order_products, 0))
        # Set shipping
        ws.cell(1, 6, self.shipping.name)
        ws.cell(1, 7, self.tracking_id)
        ws.cell(2, 6, self.country.name)
        # Set packaging
        ws.cell(11, 7, self.shipping_box_weight)

        # Set order product lines
        op_shipping = {}
        row = 11
        for suborder in self.suborders:
            if len(suborder.get_order_products()) == 0:
                continue
            row += 1
            suborder_row = row
            ws.merge_cells(f"B{row}:C{row}")
            for cell in ws[f'A{row}:M{row}'][0]:
                cell.fill = suborder_fill
            ws.cell(row, 2, f'{suborder.subcustomer.username}: {suborder.subcustomer.name}')
            ws.cell(row, 6, suborder.get_subtotal())
            ws.cell(row, 7, suborder.get_total_weight())
            # ws.cell(row, 8, suborder_shipping)
            ws.cell(row, 9, f'=F{row} + H{row}')
            ws.cell(row, 10, f'=I{row} / $E$8')
            ws.cell(row, 11, f'=I{row} / $E$9')
            ws.cell(row, 13, suborder.get_total_points())
            for op in suborder.get_order_products():
                row += 1
                op_shipping[row] = self._get_shipping_per_product(op)
                ws.cell(row, 1, op.product_id)
                ws.cell(row, 2, op.product.name_english)
                ws.cell(row, 3, op.product.name_russian)
                ws.cell(row, 4, op.quantity)
                ws.cell(row, 5, op.price)
                ws.cell(row, 6, op.price * op.quantity)
                ws.cell(row, 7, op.product.weight * op.quantity)
                ws.cell(row, 8, op_shipping[row])
                ws.cell(row, 9, ws.cell(row, 6).value + ws.cell(row, 8).value) # type: ignore
                ws.cell(row, 10, f'=I{row} / $E$8')
                ws.cell(row, 11, f'=I{row} / $E$9')
                ws.cell(row, 12, op.product.points)
                ws.cell(row, 13, op.product.points * op.quantity)
            if suborder.local_shipping != 0:
                row += 1
                ws.cell(row, 2, "Local shipping")
                ws.cell(row, 4, 1)
                ws.cell(row, 5, 2500)
                ws.cell(row, 6, 2500)
            ws.cell(suborder_row, 8, f'=SUM(H{suborder_row + 1}:H{row})')
        #TODO: Modify compensation to take into account attached orders
        # # Compensate rounding error
        if len(op_shipping) > 0 \
            and self.attached_order is None and self.attached_orders.count() == 0: #type: ignore
            diff = self.shipping_krw - \
                reduce(lambda acc, op_ship: acc + op_ship[1], op_shipping.items(), 0)
            if op_shipping.get(row) is None:
                row -= 1
            ws.cell(row, 8).value += diff
            ws.cell(row, 9).value += diff
        
        file = NamedTemporaryFile()
        order_wb.save(file.name)
        file.seek(0)
        return file

    def get_customs_label(self) -> tuple[_TemporaryFileWrapper, str]:
        '''Generates a customs label. Returns a temporary file object
        
        :returns tuple[_TemporaryFileWrapper, str]: tuple of label file 
            and file extension'''
        return self.shipping.get_customs_label(self)

    def _get_shipping_per_product(self, op):
        from app.shipping.models.shipping import PostponeShipping
        if isinstance(self.shipping, PostponeShipping):
            return self.attached_order._get_shipping_per_product(op)

        return round(self.shipping_krw / self.total_weight *
            op.product.weight * op.quantity) \
                if self.total_weight > 0 else \
                    round(self.shipping_krw / self.total_krw *
                        op.price * op.quantity)

class OrderParam(db.Model, BaseModel): # type: ignore
    '''Additional Order parameter'''
    __tablename__ = 'order_params'

    order_id = Column(String(16), ForeignKey(ORDER_ID))
    order = relationship('Order',
        backref=backref('order_params',
            collection_class=attribute_mapped_collection('name'),
            cascade='all, delete-orphan')
    )
    name = Column(String(128))
    value = Column(String(256))
