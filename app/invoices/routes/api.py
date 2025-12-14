from datetime import datetime
from functools import reduce
import os.path
import re
from tempfile import NamedTemporaryFile
from typing import Any

from more_itertools import map_reduce
import openpyxl

from flask import abort, current_app, jsonify, request
from flask.wrappers import Response
from flask_security import current_user, login_required, roles_required

from sqlalchemy import or_

from app import db
from app.currencies.models.currency import Currency
from app.invoices import bp_api_admin, bp_api_user
from app.invoices.models.invoice import Invoice
from app.invoices.models.invoice_item import InvoiceItem
from app.invoices.validators.invoice import InvoiceValidator
from app.orders.models.order import Order
from app.tools import modify_object, prepare_datatables_query, stream_and_close


@bp_api_user.route("/new", methods=["POST"])
@login_required
def create_invoice():
    """Creates invoice for provided orders"""
    with InvoiceValidator(request) as validator:
        if not validator.validate():
            return Response(f"Couldn't create an Invoice\n{validator.errors}", status=409)

    payload: dict[str, Any] = request.get_json() # type: ignore
    if not current_user.has_role('admin'):
        orders = Order.query.filter(Order.id.in_(payload["order_ids"]), Order.user == current_user).all()
    else:
        orders = Order.query.filter(Order.id.in_(payload["order_ids"])).all()
    currency = Currency.query.get(payload["currency"])
    rate = currency.get_rate()
    invoice = Invoice(currency_code=currency.code, user=current_user)
    # invoice_items = []
    invoice.when_created = datetime.now()  # type: ignore
    cumulative_order_products = map_reduce(
        [order_product for order in orders for order_product in order.order_products],
        keyfunc=lambda ii: (
            ii.product_id,
            ii.product.name_english if ii.product.name_english else ii.product.name,
            ii.price,
        ),
        valuefunc=lambda op: op.quantity,
        reducefunc=sum,
    )
    for order in orders:
        order.invoice = invoice

    db.session.add(invoice) #type: ignore
    db.session.add_all( #type: ignore
        [
            InvoiceItem(
                invoice=invoice,
                product_id=op[0][0],
                price=round(op[0][2] * rate, 2),
                quantity=op[1],
            )
            for op in cumulative_order_products.items()
        ]
    )
    db.session.add( #type: ignore
        InvoiceItem(
            invoice=invoice,
            product_id="SHIPPING",
            price=round(
                reduce(lambda acc, o: acc + o.shipping_krw * rate, orders, 0), 2
            ),
            quantity=1,
        )
    )
    db.session.commit() #type: ignore
    return jsonify({"status": "success", "invoice_id": invoice.id})


@bp_api_user.route("/", defaults={"invoice_id": None}, strict_slashes=False)
@bp_api_user.route("/<invoice_id>")
@login_required
def get_invoices(invoice_id):
    """
    Returns all or selected invoices in JSON:
    """

    invoices = Invoice.query
    if not current_user.has_role('admin'):
        invoices = invoices.filter_by(user=current_user)

    if invoice_id is not None:
        invoices = invoices.filter_by(id=invoice_id)
    else:  # here we check whether request is filtered by DataTables
        if request.values.get("draw") is not None:
            invoices, records_total, records_filtered = prepare_datatables_query(
                invoices,
                request.values,
                or_(
                    Invoice.id.like(f"%{request.values['search[value]']}%"),  # type: ignore
                    Invoice.orders.any(
                        Order.id.like(f"%{request.values['search[value]']}%")
                    ),
                    Invoice.customer.like(f"%{request.values['search[value]']}%"),
                ),
            )
            return jsonify(
                {
                    "draw": request.values["draw"],
                    "recordsTotal": records_total,
                    "recordsFiltered": records_filtered,
                    "data": [entry.to_dict() for entry in invoices],
                }
            )
        else:  # By default we return only 100 invoices
            invoices = invoices.limit(10)

    return jsonify([entry.to_dict(details=invoices.count() == 1) for entry in invoices])


def get_invoice_order_products(invoice: Invoice):
    cumulative_order_products = map_reduce(
        invoice.get_invoice_items(),
        keyfunc=lambda ii: (
            ii.product_id,
            ii.product.name_english if ii.product.name_english else ii.product.name,
            ii.product.name_russian if ii.product.name_russian else ii.product.name,
            ii.product.weight,
            ii.price,
        ),
        valuefunc=lambda op: op.quantity,
        reducefunc=sum,
    )

    result = list(
        map(
            lambda ii: {
                "id": ii[0][0],
                "name": ii[0][1],
                'name_russian': ii[0][2],
                "price": ii[0][4],
                "quantity": ii[1],
                "weight": ii[0][3],
                "subtotal": ii[0][4] * ii[1],
            },
            cumulative_order_products.items(),
        )
    )
    return result


@bp_api_user.route("/template")
@login_required
def get_templates():
    global_template_path = os.path.dirname(__file__) + "/../templates"
    local_template_path = current_app.config["INVOICE_TEMPLATES_PATH"]
    templates = [
        "[global]/" + file_name
        for file_name in os.listdir(global_template_path)
        if ".xlsx" in file_name
    ] + (
        [
            "[local]/" + file_name
            for file_name in os.listdir(local_template_path)
            if ".xlsx" in file_name
        ]
        if os.path.exists(local_template_path)
        else []
    )
    return jsonify(templates)


def create_invoice_excel(reference_invoice: Invoice, template: str):
    def render_cell(cell, template, vars):
        if isinstance(cell, openpyxl.cell.cell.MergedCell): #type: ignore
            return
        params = re.findall("(?<={{).+?(?=}})", str(template))
        for param in params:
            template = template.replace(
                f"{{{{{param}}}}}",
                str(eval(param, vars)),
            )
        cell.value = template

    def render_sheet(sheet, rows, cols):
        for row in rows:
            for col in cols:
                render_cell(
                    sheet.cell(row, col),
                    sheet.cell(row, col).value,
                    {
                        "reference_invoice": reference_invoice,
                        "payee": payee,
                        "suffix": suffix,
                        "invoice_dict": invoice_dict,
                        "total": total,
                        'pieces': pieces,
                        'currency_code': reference_invoice.currency_code,
                    },
                )

    def render_items(sheet, list, start_row, cols):
        row = start_row
        templates = [sheet.cell(row, col).value for col in range(1, cols + 1)]
        for item in list:
            for col in range(0, cols):
                render_cell(
                    sheet.cell(row, col + 1),
                    templates[col],
                    {"op": item},
                )
            row += 1

    global_path = os.path.dirname(__file__) + "/../templates"
    local_path = current_app.config["INVOICE_TEMPLATES_PATH"]
    if template is None:
        template = global_path + "/default.xlsx"
    template = template.replace("[global]", global_path).replace("[local]", local_path)
    invoice_wb = openpyxl.open(template)
    invoice_dict = reference_invoice.to_dict(details=True)
    order_products = get_invoice_order_products(reference_invoice)
    total = reduce(lambda acc, op: acc + op["subtotal"], order_products, 0)
    pieces = reduce(lambda qty, op: qty + op['quantity'], order_products, 0)
    payee = reference_invoice.orders[0].get_payee()
    suffix = "ES" if len(reference_invoice.orders) > 1 else ""
    ws = invoice_wb.worksheets[0]
    pl = invoice_wb.worksheets[1]
    render_sheet(
        ws, rows=list(range(1, 26)) + list(range(305, 313)), cols=list(range(1, 6))
    )

    render_sheet(
        pl, rows=list(range(1, 26)) + list(range(305, 313)), cols=list(range(1, 6))
    )

    # Set order product lines
    row = 31
    last_row = 304

    render_items(ws, order_products, row, 6)
    ws.delete_rows(
        row + len(order_products), last_row - row - len(order_products) + 1
    )
    render_items(pl, order_products, row, 4)
    pl.delete_rows(
        row + len(order_products), last_row - row - len(order_products) + 1
    )

    file = NamedTemporaryFile()
    invoice_wb.save(file.name)
    file.seek(0)
    return file


@bp_api_admin.route("/<invoice_id>", methods=["POST"])
@roles_required("admin")
def save_invoice(invoice_id):
    invoice = Invoice.query.get(invoice_id)
    if not invoice:
        abort(Response(f"The invoice <{invoice_id}> was not found", status=404))
    payload = request.get_json()
    if not payload:
        abort(Response("No invoice data was provided", status=400))
    modify_object(invoice, payload, ["customer", "export_id", "payee"])
    db.session.commit() #type: ignore
    return jsonify({"data": [invoice.to_dict()]})


@bp_api_user.route("/<invoice_id>/excel")
@login_required
def get_invoice_excel(invoice_id):
    """
    Generates an Excel file for an invoice
    """
    invoice = Invoice.query.get(invoice_id)
    if not invoice:
        abort(Response(f"The invoice <{invoice_id}> was not found", status=404))
    if not current_user.has_role('admin') and invoice.user != current_user:
        abort(Response(f"The invoice <{invoice_id}> does not belong to you", status=403))
    if invoice.invoice_items_count == 0:
        abort(Response(f"The invoice <{invoice_id}> has no items", status=406))

    file = create_invoice_excel(
        reference_invoice=invoice, template=request.args.get("template") #type: ignore
    )
    return current_app.response_class(
        stream_and_close(file),
        headers={
            "Content-Disposition": f'attachment; filename="{invoice_id}.xlsx"',
            "Content-Type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        },
    )


@bp_api_user.route("/excel")
@login_required
def get_invoice_cumulative_excel():
    """
    Returns cumulative Excel of several invoices
    invoice IDs are provides as URL arguments
    Resulting Excel isn't saved anywhere
    """
    cumulative_invoice = Invoice()
    valid_invoices = []
    for invoice_id in request.args.getlist("invoices"):
        invoice = Invoice.query.get(invoice_id)
        if invoice and (current_user.has_role('admin') or invoice.user == current_user):
            valid_invoices.append(invoice)
            if not cumulative_invoice.customer:
                cumulative_invoice.customer = invoice.customer
            if not cumulative_invoice.payee and invoice.payee:
                cumulative_invoice.payee = invoice.payee
            cumulative_invoice.orders += invoice.orders

    file = create_invoice_excel(
        reference_invoice=cumulative_invoice, template=request.args.get("template") #type: ignore
    )
    return current_app.response_class(
        stream_and_close(file),
        headers={
            "Content-Disposition": 'attachment; filename="cumulative_invoice.xlsx"',
            "Content-Type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        },
    )


@bp_api_admin.route("/<invoice_id>/item/<invoice_item_id>", methods=["POST"])
@roles_required("admin")
def save_invoice_item(invoice_id, invoice_item_id):
    """
    Creates or modifies existing invoice item
    """
    payload = request.get_json()
    if not payload:
        abort(Response("No data was provided", status=400))
    invoice = Invoice.query.get(invoice_id)
    invoice_item = None
    if not invoice:
        abort(Response(f"No invoice <{invoice_id}> was found", status=404))
    if invoice_item_id != "new":
        invoice_item = invoice.get_invoice_items().filter_by(id=invoice_item_id).first()
        if not invoice_item:
            abort(
                Response(f"No invoice item <{invoice_item_id}> was found", status=404)
            )
    else:
        invoice_item = InvoiceItem(invoice=invoice, when_created=datetime.now())

    modify_object(invoice_item, payload, ["product_id", "price", "quantity"])
    if invoice_item_id == "new":
        db.session.add(invoice_item) #type: ignore
    db.session.commit() #type: ignore
    return jsonify(invoice_item.to_dict())


@bp_api_admin.route("/<invoice_id>/item/<invoice_item_id>", methods=["DELETE"])
@roles_required("admin")
def delete_invoice_item(invoice_id, invoice_item_id):
    """
    Deletes existing invoice item
    """
    invoice = Invoice.query.get(invoice_id)
    if not invoice:
        abort(Response(f"No invoice <{invoice_id}> was found", status=404))
    invoice_item = invoice.get_invoice_items().filter_by(id=invoice_item_id).first()
    if not invoice_item:
        abort(Response(f"No invoice item<{invoice_item_id}> was found", status=404))
    db.session.delete(invoice_item) #type: ignore
    db.session.commit() #type: ignore
    return jsonify({"status": "success"})
