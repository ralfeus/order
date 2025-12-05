from __future__ import annotations
from datetime import datetime
import logging
import os.path
from tempfile import _TemporaryFileWrapper, NamedTemporaryFile

import openpyxl

import app.orders.models as o
from app.shipping.models.shipping import Shipping
from exceptions import OrderError

class Cargo(Shipping):
    __mapper_args__ = {'polymorphic_identity': 'cargo'} #type: ignore

    def get_edit_url(self):
        from .. import bp_client_admin
        return f"{bp_client_admin.url_prefix}/{self.id}"

    def get_shipping_cost(self, destination, weight):
        '''
        Returns shipping cost for cargo - always 0 if country is allowed
        '''
        if isinstance(destination, str):
            destination_id = destination
        else:
            destination_id = destination.id
        rate = self.rates.filter_by(destination=destination_id).first()
        if rate:
            return 0
        from exceptions import NoShippingRateError
        raise NoShippingRateError()

    def get_customs_label(self, order: 'o.Order') -> tuple[_TemporaryFileWrapper, str]: #type: ignore
        '''Generates an invoice in excel format. 
        
        :returns tuple[_TemporaryFileWrapper, str]: a temporary file object 
            and file extension (xlsx)'''
        logger = logging.getLogger('Cargo.get_customs_label')
        if len(order.order_products) == 0:
            raise OrderError("The order has no products")
        # suborder_fill = PatternFill(
        #     start_color='00FFFF00', end_color='00FFFF00', fill_type='solid')
        country = order.country.id
        try:
            method = self.__getattribute__(f'_Cargo__get_customs_label_for_{country}')
            return method(order)
        except AttributeError:
            return self.__get_customs_label_generic(order) #type: ignore

    def __get_customs_label_generic(self, order: 'o.Order'): #type: ignore
        return None, None

    def __get_customs_label_for_uz(self, order: 'o.Order'): #type: ignore
        package_path = os.path.dirname(__file__) + '/..'
        wb = openpyxl.open(f'{package_path}/templates/customs_label_uz.xlsx')
        ws = wb.worksheets[0]
        ws.cell(3, 6, order.id)
        ws.cell(7, 4, order.customer_name)
        ws.cell(8, 4, order.address)
        ws.cell(9, 4, order.params.get('shipping.passport_number'))
        ws.cell(9, 6, order.params.get('shipping.tax_id'))
        ws.cell(10, 4, order.phone)
        ws.cell(82, 4, datetime.date(datetime.now()))
        row = 13
        total_quantity = total_weight = total_amount = 0
        currency_rate = order.total_krw / order.total_usd
        for op in order.order_products:
            ws.cell(row, 3, op.product.name_english)
            ws.cell(row, 4, op.quantity)
            ws.cell(row, 6, op.product.weight * op.quantity / 1000)
            ws.cell(row, 7, op.price / currency_rate)
            ws.cell(row, 8, op.price * op.quantity / currency_rate)
            total_quantity += op.quantity
            total_weight += op.product.weight * op.quantity
            total_amount += op.price * op.quantity / currency_rate
            row += 1
        ws.cell(79, 4, total_quantity)
        ws.cell(79, 6, f"{total_weight / 1000} kg")
        ws.cell(79, 8, total_amount)
        last_row =  78
        ws.delete_rows(row, last_row - row + 1)
        file = NamedTemporaryFile()
        wb.save(file.name)
        return file, 'xlsx'
